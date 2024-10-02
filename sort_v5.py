import sys
import subprocess
import os
import re
import json
from datetime import datetime
import glob
import shutil
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import win32com.client
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, scrolledtext
import threading
import pythoncom
import queue
import uuid
import ctypes
import requests
# Скрываем консольное окно
if sys.platform.startswith('win'):
    ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

def install_dependencies():
    dependencies = [
        'pillow',
        'pywin32',
        'pyperclip',
        'requests',
        'openpyxl',
        'python-docx'
    ]
    
    print("Установка зависимостей...")
    for package in dependencies:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"Успешно установлен пакет: {package}")
        except subprocess.CalledProcessError:
            print(f"Не удалось установить пакет: {package}")
    print("Установка зависимостей завершена.")

try:
    # Попытка импорта необходимых модулей
    import requests
    from PIL import Image
    import win32com.client
    import pyperclip
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    print("Попытка установки недостающих зависимостей...")
    install_dependencies()
    # Повторная попытка импорта после установки
    import requests
    from PIL import Image
    import win32com.client
    import pyperclip

#  Добавляем версии
SCRIPT_VERSION = "6.0.0"
DOCM_VERSION = "1.0.6"
EXCEL_TEMPLATE_VERSION = "1.0.1"  # Добавляем версию для шаблона Excel

GITHUB_REPO = "cyberkek587/drain_tg"
SCRIPT_NAME = "sort_v5.py"
DOCM_NAME = "word_jpg_auto_v5.docm"
EXCEL_TEMPLATE_NAME = "excel_summary_template.xlsm"

def check_for_updates():
    """Проверяет наличие обновлений скрипта, .docm файла и шаблона Excel на GitHub."""
    try:
        requests.get("https://github.com", timeout=5)
    except requests.ConnectionError:
        print("Нет подключения к интернету. Пропускаем проверку обновлений.")
        return False

    try:
        script_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{SCRIPT_NAME}"
        docm_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{DOCM_NAME}"
        excel_template_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{EXCEL_TEMPLATE_NAME}"
        
        script_response = requests.get(script_url, timeout=10)
        docm_response = requests.get(docm_url, timeout=10)
        excel_template_response = requests.get(excel_template_url, timeout=10)
        
        update_available = False
        
        if script_response.status_code == 200:
            remote_script = script_response.text
            remote_script_version = re.search(r'SCRIPT_VERSION\s*=\s*"([\d.]+)"', remote_script)
            if remote_script_version and remote_script_version.group(1) > SCRIPT_VERSION:
                print(f"Доступно обновление скрипта. Текущая версия: {SCRIPT_VERSION}, новая версия: {remote_script_version.group(1)}")
                update_available = True
        
        if docm_response.status_code == 200:
            remote_docm = docm_response.content
            remote_docm_version = re.search(r'DOCM_VERSION\s*=\s*"([\d.]+)"', remote_script)
            if remote_docm_version and remote_docm_version.group(1) > DOCM_VERSION:
                print(f"Доступно обновление .docm файла. Текущая версия: {DOCM_VERSION}, новая версия: {remote_docm_version.group(1)}")
                update_available = True
        
        if excel_template_response.status_code == 200:
            remote_excel_template = excel_template_response.content
            remote_excel_template_version = re.search(r'EXCEL_TEMPLATE_VERSION\s*=\s*"([\d.]+)"', remote_script)
            if remote_excel_template_version and remote_excel_template_version.group(1) > EXCEL_TEMPLATE_VERSION:
                print(f"Доступно обновление шаблона Excel. Текущая версия: {EXCEL_TEMPLATE_VERSION}, новая версия: {remote_excel_template_version.group(1)}")
                update_available = True
        
        if update_available:
            choice = input("Хотите обновить файлы? (д/н): ").lower()
            if choice == 'д':
                print("Обновляем файлы...")
                try:
                    # Создаем резервную копию текущего скрипта
                    backup_name = f"{SCRIPT_NAME}.bak"
                    shutil.copy(__file__, backup_name)
                    print(f"Создана резервная копия скрипта: {backup_name}")

                    with open(__file__, 'w', encoding='utf-8') as file:
                        file.write(remote_script)
                    with open(DOCM_NAME, 'wb') as file:
                        file.write(remote_docm)
                    
                    # Скачиваем шаблон Excel
                    excel_template_path = os.path.join(os.path.dirname(__file__), EXCEL_TEMPLATE_NAME)
                    with open(excel_template_path, 'wb') as file:
                        file.write(excel_template_response.content)
                    print("Шаблон Excel обновлен.")
                    
                    print("Файлы обновлены. Перезапустите программу.")
                    input("Нажмите Enter для завершения...")
                    return True
                except Exception as e:
                    print(f"Ошибка при обновлении файлов: {e}")
                    print("Восстанавливаем предыдущую версию скрипта...")
                    shutil.copy(backup_name, __file__)
                    print("Предыдущая версия скрипта восстановлена.")
                    print("Продолжаем работу с текущими версиями.")
                    input("Нажмите Enter для продолжения...")
        else:
            print(f"У вас установлены последние версии скрипта ({SCRIPT_VERSION}), .docm файла ({DOCM_VERSION}) и шаблона Excel ({EXCEL_TEMPLATE_VERSION}).")
        
        return False
    except Exception as e:
        print(f"Ошибка при проверке обновлений: {e}")
        print("Продолжаем работу с текущими версиями файлов.")
        input("Нажмите Enter для продолжения...")
        return False

# Вызываем функцию проверки обновлений
if check_for_updates():
    sys.exit()

# Проверяем наличие файлов и скачиваем их, если отсутствуют
def check_and_download_file(file_name, url):
    if not os.path.exists(file_name):
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                with open(file_name, 'wb') as file:
                    file.write(response.content)
                print(f"Файл {file_name} успешно скачан.")
            else:
                print(f"Не удалось скачать файл {file_name}. Код ответа: {response.status_code}")
        except Exception as e:
            print(f"Ошибка при скачивании файла {file_name}: {e}")

check_and_download_file(DOCM_NAME, f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{DOCM_NAME}")
check_and_download_file(EXCEL_TEMPLATE_NAME, f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{EXCEL_TEMPLATE_NAME}")

# Словарь для хранения тем
THEMES = {
    2: "Ручная чистка",
    4: "Перехваты Авраменко Святогеоргиевска",
    6: "Ежедневные отчеты",
    9: "АВАРИИ",
    12: "Отчет о выполнении работ по объектам",
    18: "Работа наемной техники",
    58: "Объект Днепропетровское шоссе",
    121: "Стройка ливневки колодцы"
}

def update_themes(filename):
    """
    Обновляет словарь THEMES новыми темами из файла JSON,
    сохраняя приоритет существующих названий для указанных ID.

    Args:
        filename: Имя файла JSON.
    """
    global THEMES
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)

    existing_themes = THEMES.copy()
    new_themes = {}
    updated_themes = []

    for message in data['messages']:
        if message['type'] == 'service' and message['action'] == 'topic_created':
            theme_id = int(message['id'])
            if theme_id not in existing_themes:
                new_themes[theme_id] = message['title']
                updated_themes.append((theme_id, message['title']))

    # Обновляем THEMES, сохраняя приоритет существующих названий
    THEMES.update(new_themes)

    #print(f"Обновлен словарь тем. Всего тем: {len(THEMES)}")
    if updated_themes:
        #print("Новые темы:")
        for id, theme in updated_themes:
            #print(f"{id}: {theme}")
            pass
    else:
        #print("Новых тем не обнаружено.")
        pass
    
    print("Текущий список тем:")
    for id, theme in THEMES.items():
        print(f"{id}: {theme}")
    print("Обновление тем завершено")

def get_date_range(messages):
    """Находит самую раннюю и самую позднюю даты сообщений."""
    dates = [int(msg['date_unixtime']) for msg in messages if 'date_unixtime' in msg]
    if not dates:
        return None, None
    earliest = datetime.fromtimestamp(min(dates))
    latest = datetime.fromtimestamp(max(dates))
    return earliest.strftime("%d-%m-%y"), latest.strftime("%d-%m-%y")

def process_json(json_file):
    """Обрабатывает JSON файл и сортирует сообщения по темам."""
    global sorted_folder_name, date_range  # Объявляем переменные как глобальные
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    messages = data['messages']
    #print("Начали обработку сообщений")

    photos_path = os.path.join(os.path.dirname(__file__), 'photos')
    
    # Получаем диапазон дат и создаем новое имя для папки sorted
    start_date, end_date = get_date_range(messages)
    date_range = f"{start_date}_to_{end_date}"
    sorted_folder_name = f"sorted_{date_range}"
    sorted_path = os.path.join(os.path.dirname(__file__), sorted_folder_name)
    #print(f"Путь к папке с отсортированными файлами: {sorted_path}")

    theme_folders = {}
    processed_photos = set()  # Множество для отслеживания обработанных фотографий

    for message in messages:
        # Пропускаем сообщения, если 'from_id' или 'reply_to_message_id' отсутствуют
        if 'from_id' not in message or 'reply_to_message_id' not in message:
            continue

        # Пропускаем сообщения с group_id = 6
        if 'group_id' in message and message['group_id'] == 6:
            continue

        # Пропускаем сообщения, если поле text больше 255 символов
        if 'text' in message and len(message['text']) > 235:
            continue

        #print(f"Обрабатываем сообщение с id: {message['id']}")
        reply_to_id = int(message['reply_to_message_id'])
        if reply_to_id in THEMES:
            theme = THEMES[reply_to_id]
            theme_path = os.path.join(sorted_path, theme.replace(" ", "_"))

            if theme_path not in theme_folders:
                theme_folders[theme_path] = {}

            # Группируем сообщения
            grouped_messages = [message]
            for msg in messages:
                # Преобразуем значения date_unixtime в числа
                msg_date_unixtime = int(msg['date_unixtime'])
                message_date_unixtime = int(message['date_unixtime'])

                # Условие группировки по разнице во времени не более 3 секунд
                if (
                        abs(msg_date_unixtime - message_date_unixtime) <= 3
                        and msg['from_id'] == message['from_id']
                        and msg['reply_to_message_id'] == message['reply_to_message_id']
                ):
                    grouped_messages.append(msg)
                    #print(f"Добавлено сообщение в группу: {msg['id']}")

            # Проверяем, есть ли в группе сообщения с текстом
            subfolder_name = "Без_текста"
            for msg in grouped_messages:
                if 'text' in msg and msg['text']:
                    # Заменяем /n на _ в тексте
                    subfolder_name = msg['text'].replace(" ", "_").replace("\n", "_")
                    # Очищаем имя папки от недопустимых символов
                    subfolder_name = re.sub(r'[^\w\s-]', '_', subfolder_name).strip()
                    break

            subfolder_path = os.path.join(theme_path, subfolder_name)
            
            # Проверяем, есть ли фотографии для перемещения
            photos_to_move = [msg for msg in grouped_messages if 'photo' in msg]
            if photos_to_move:
                if subfolder_path not in theme_folders[theme_path]:
                    theme_folders[theme_path][subfolder_path] = []
                theme_folders[theme_path][subfolder_path].extend(photos_to_move)

    # Создаем папки и копируем фотографии
    for theme_path, subfolders in theme_folders.items():
        if subfolders:
            os.makedirs(theme_path, exist_ok=True)
            #print(f"Создана папка темы: {theme_path}")
            
            for subfolder_path, photos in subfolders.items():
                os.makedirs(subfolder_path, exist_ok=True)
                #print(f"Создана подпапка: {subfolder_path}")
                
                for msg in photos:
                    source_file = os.path.join(photos_path, os.path.basename(msg['photo']))
                    target_file = os.path.join(subfolder_path, os.path.basename(msg['photo']))

                    if os.path.exists(source_file):
                        shutil.copy2(source_file, target_file)
                        #print(f"Скопирован файл: {source_file} в {target_file}")
                        processed_photos.add(source_file)
                    else:
                        #print(f"Файл не найден: {source_file}")
                        pass

    # Удаляем обработанные фотографии из папки photos
    for photo in processed_photos:
        try:
            os.remove(photo)
            #print(f"Удален файл: {photo}")
        except OSError as e:
            #print(f"Ошибка при удалении файла {photo}: {e}")
            pass

    # Проверяем, существует ли папка photos, и удаляем ее, если она пуста
    if os.path.exists(photos_path):
        if not os.listdir(photos_path):
            os.rmdir(photos_path)
            #print("Папка photos удалена, так как она пуста.")
        else:
            #print("ВНИМАНИЕ: В папке photos остались нерассортированные фотографии.")
            pass
    else:
        #print("Папка photos не существует.")
        pass

    clear_terminal()
    #print("Обработка сообщений завершена")
    #print(f"Всего обработано уникальных фотографий: {len(processed_photos)}")
    #print("Обработка JSON завершена")

def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')

def parse_folder_selection(input_string, max_index):
    selected_indices = set()
    parts = input_string.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            start, end = map(int, part.split('-'))
            selected_indices.update(range(start - 1, min(end, max_index)))
        else:
            try:
                index = int(part) - 1
                if 0 <= index < max_index:
                    selected_indices.add(index)
            except ValueError:
                pass
    return sorted(list(selected_indices))

def sanitize_folder_name(name):
    """Очищает имя папки от недопустимых символов."""
    # Запрещенные символы в Windows
    forbidden_chars = r'<>:"/\|?*'
    # Заменяем запрещенные символы на подчеркивание
    for char in forbidden_chars:
        name = name.replace(char, '_')
    # Удаляем точки и пробелы в начале и конце имени
    name = name.strip('. ')
    # Ограничиваем длину имени папки до 255 символов
    return name[:255]

def update_excel(docx_files, prefix):
    template_path = os.path.join(os.path.dirname(__file__), EXCEL_TEMPLATE_NAME)
    if not os.path.exists(template_path):
        print("Шаблон Excel не найден. Пожалуйста, запустите проверку обновлений.")
        return

    docx_folder = os.path.join(os.path.dirname(__file__), f'docx_{date_range}')
    os.makedirs(docx_folder, exist_ok=True)
    excel_path = os.path.join(docx_folder, 'Содержание.xlsm')
    
    if not os.path.exists(excel_path):
        shutil.copy(template_path, excel_path)
        #print(f"Создан новый файл {excel_path} на основе шаблона.")
    
    # Открываем существующий файл Excel
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)

    sheet_name = prefix if prefix else "Без префикса"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["№", "Название объекта", "Количество", "Примечание", "Печать"])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = Font(bold=True)
    else:
        ws = wb[sheet_name]

    # Определяем стили для границ
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Удаляем существующую итоговую строку, если она есть
    if ws.cell(row=ws.max_row, column=2).value == "Всего":
        ws.delete_rows(ws.max_row)

    # Добавляем новые данные
    for docx_file in docx_files:
        file_name = os.path.splitext(os.path.basename(docx_file))[0]
        # Удаляем префикс из отображаемого текста
        display_name = file_name[len(prefix)+1:] if prefix and file_name.startswith(prefix) else file_name
        relative_path = os.path.relpath(docx_file, start=os.path.dirname(excel_path))
        ws.append(["", f'=HYPERLINK("{relative_path}", "{display_name}")', "", "", ""])

    # Добавляем итоговую строку
    ws.append(["", "Всего", "", "", ""])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # Нумеруем непустые строки и определяем диапазон суммирования
    row_number = 1
    first_data_row = None
    last_data_row = None
    for row in range(2, ws.max_row):
        if ws.cell(row=row, column=2).value and ws.cell(row=row, column=2).value != "Всего":
            ws.cell(row=row, column=1).value = row_number
            row_number += 1
            if first_data_row is None:
                first_data_row = row
            last_data_row = row

    # Обновляем формулу в итоговой строке
    if first_data_row and last_data_row:
        ws.cell(row=ws.max_row, column=3).value = f"=SUM(C{first_data_row}:C{last_data_row})"

    # Применяем границы и выравнивание ко всем ячейкам
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Применяем автоподбор ширины столбцов
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.row == 1 or cell.row == ws.max_row:  # Заголовок или итоговая строка
                max_length = max(max_length, len(str(cell.value)))
            elif column == 2:  # Второй столбец (с гиперссылками)
                if cell.value and cell.value.startswith('=HYPERLINK'):
                    # Извлекаем текст гиперссылки
                    display_text = cell.value.split('"')[3]
                    max_length = max(max_length, len(display_text))
            else:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 4
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Добавляем "кнопку" для печати
    button_cell = ws.cell(row=1, column=6)
    button_cell.value = "ПЕЧАТЬ"
    button_cell.font = Font(bold=True, color="FFFFFF")
    button_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    button_cell.alignment = Alignment(horizontal='center', vertical='center')
    button_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Устанавливаем ширину столбца для кнопки
    ws.column_dimensions['F'].width = 15

    # Сохраняем изменения в файле Excel
    wb.save(excel_path)
    #print(f"Файл Excel обновлен: {excel_path}")

    # Добавляем макрос к кнопке с помощью win32com
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    wb = xl.Workbooks.Open(excel_path)
    ws = wb.Worksheets(sheet_name)
    
    # Добавляем кнопку и привязываем к ней макрос
    button = ws.Buttons().Add(ws.Range("F1").Left, ws.Range("F1").Top, ws.Range("F1").Width, ws.Range("F1").Height)
    button.OnAction = "ПечатьДокументов"
    button.Caption = "ПЕЧАТЬ"
    button.Name = "ПечатьДокументов"

    wb.Save()
    xl.Quit()

    #print("Кнопка для запуска макроса печати добавлена.")

def process_photos_and_create_docx(folder_path, folder_name, prefix=""):
    """Обрабатывает фотографии и создает docx файл с добавлением префикса к имени файла."""
    word = win32com.client.Dispatch("Word.Application")
    status = ""

    # Используем имя папки в качестве текста
    text_value = f"{prefix}_{folder_name}" if prefix else folder_name

    # Копируем путь к папке и имя файла в буфер обмена, разделяя их символом |
    pyperclip.copy(f"{folder_path}|{text_value}")

    image_files = glob.glob(os.path.join(folder_path, '*.jpg'))
    image_files.sort(key=lambda x: os.path.getctime(x))
    total_images = len(image_files)

    print("Обработка изображений:")
    for file_path in image_files:
        try:
            with Image.open(file_path) as img:
                width, height = img.size
                if width > height:
                    img = img.transpose(method=Image.ROTATE_90)
                    img.save(file_path)
        except Exception as e:
            status += f"Не удалось обработать изображение: {os.path.basename(file_path)}. Ошибка: {str(e)}\n"

    status += f"Обработано {total_images} изображений.\n"

    script_dir = os.path.dirname(os.path.abspath(__file__))
    docm_file_path = os.path.join(script_dir, 'word_jpg_auto_v5.docm')

    if not os.path.exists(docm_file_path):
        return "Ошибка: файл word_jpg_auto_v5.docm не найден"

    print("Создание docx файла...")
    word.Documents.Open(docm_file_path)
    word.Application.Run("Макрос1")
    word.ActiveDocument.Close()
    word.Quit()

    docx_folder = os.path.join(script_dir, f'docx_{date_range}')
    os.makedirs(docx_folder, exist_ok=True)

    docx_files = []
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".docx"):
            source_path = os.path.join(folder_path, file_name)
            new_file_name = f"{prefix}_{file_name}" if prefix and not file_name.startswith(prefix) else file_name
            destination_path = os.path.join(docx_folder, new_file_name)
            shutil.move(source_path, destination_path)
            pyperclip.copy(destination_path)
            status += f"Сформирован {new_file_name} и перемещен в папку docx\n"
            docx_files.append(destination_path)

    update_excel(docx_files, prefix)

    return status

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Обработка фотографий и создание документов")
        self.geometry("1000x600")

        self.create_widgets()
        
        self.queue = queue.Queue()
        self.task_queue = queue.Queue()
        self.processing = False
        self.total_tasks = 0
        self.completed_tasks = 0
        self.is_processing_docx = False
        self.after(100, self.process_queue)
        self.after(100, self.process_task_queue)
        self.after(1000, self.update_summary_button_state)
        self.folder_ids = {}  # Словарь для хранения уникальных идентификаторов папок

    def create_widgets(self):
        # Создаем фрейм для разделения дерева и консоли
        self.main_frame = ttk.PanedWindow(self, orient=tk.VERTICAL)
        self.main_frame.pack(expand=True, fill='both')

        # Фрейм для дерева с полосой прокрутки
        self.tree_frame = ttk.Frame(self.main_frame)
        self.main_frame.add(self.tree_frame, weight=4)

        # Создаем полосу прокрутки для дерева
        self.tree_scrollbar = ttk.Scrollbar(self.tree_frame)
        self.tree_scrollbar.pack(side='right', fill='y')

        self.tree = ttk.Treeview(self.tree_frame, yscrollcommand=self.tree_scrollbar.set)
        self.tree.pack(expand=True, fill='both')

        # Привязываем полосу прокрутки к дереву
        self.tree_scrollbar.config(command=self.tree.yview)

        # Привязываем двойной щелчок к функции переименования
        self.tree.bind("<Double-1>", self.on_double_click)

        # Добавляем тег для зеленой заливки
        self.tree.tag_configure('green', background='lightgreen')

        # Добавляем тег для серой заливки
        self.tree.tag_configure('gray', background='lightgray')

        self.button_frame = ttk.Frame(self.tree_frame)
        self.button_frame.pack(fill='x', padx=5, pady=5)

        self.merge_button = ttk.Button(self.button_frame, text="Объединить выбранные", command=self.merge_selected)
        self.merge_button.pack(side='left', padx=5)

        self.find_replace_button = ttk.Button(self.button_frame, text="Найти и заменить", command=self.find_and_replace)
        self.find_replace_button.pack(side='left', padx=5)

        self.create_docx_button = ttk.Button(self.button_frame, text="Создать DOCX", command=self.create_docx)
        self.create_docx_button.pack(side='left', padx=5)

        self.open_summary_button = ttk.Button(self.button_frame, text="Открыть содержание", command=self.open_summary)
        self.open_summary_button.pack(side='left', padx=5)
        self.open_summary_button.config(state='disabled')  # Изначально кнопка неактивна

        # Добавляем индикатор прогресса
        self.progress = ttk.Progressbar(self.button_frame, orient="horizontal", length=200, mode="determinate")
        self.progress.pack(side='left', padx=5)

        # Добавляем метку для отображения прогресса
        self.progress_label = ttk.Label(self.button_frame, text="0%")
        self.progress_label.pack(side='left', padx=5)

        # Создаем текстовое поле для вывода (консоль)
        self.console_frame = ttk.Frame(self.main_frame)
        self.main_frame.add(self.console_frame, weight=1)  # Уменьшаем вес консоли

        self.console = scrolledtext.ScrolledText(self.console_frame, wrap=tk.WORD, height=10)
        self.console.pack(expand=True, fill='both')

        # Добавьте эту строку в метод create_widgets
        self.status_label = ttk.Label(self, text="")
        self.status_label.pack(side='bottom', fill='x')
    
    def find_and_replace(self):
            selected_items = self.tree.selection()
            if not selected_items:
                messagebox.showwarning("Предупреждение", "Выберите папки для поиска и замены.")
                return

            # Создаем новое окно для ввода текста
            dialog = tk.Toplevel(self)
            dialog.title("Найти и заменить")
            dialog.geometry("300x180")

            ttk.Label(dialog, text="Найти:").pack(pady=5)
            find_entry = ttk.Entry(dialog, width=40)
            find_entry.pack(pady=5)

            ttk.Label(dialog, text="Заменить на:").pack(pady=5)
            replace_entry = ttk.Entry(dialog, width=40)
            replace_entry.pack(pady=5)

            def perform_replace():
                find_text = find_entry.get()
                replace_text = replace_entry.get()

                if not find_text:
                    messagebox.showwarning("Предупреждение", "Введите текст для поиска.")
                    return

                renamed_count = 0
                for item in selected_items:
                    old_name = self.tree.item(item)['text']
                    new_name = old_name.replace(find_text, replace_text)
                    
                    if new_name != old_name:
                        parent = self.tree.parent(item)
                        if parent:  # Это подпапка
                            theme_path = os.path.join(os.path.dirname(__file__), sorted_folder_name, self.tree.item(parent)['text'])
                        else:  # Это тема
                            theme_path = os.path.join(os.path.dirname(__file__), sorted_folder_name)
                        
                        old_path = os.path.join(theme_path, old_name)
                        new_path = os.path.join(theme_path, new_name)

                        try:
                            os.rename(old_path, new_path)
                            self.tree.item(item, text=new_name)
                            renamed_count += 1
                        except OSError as e:
                            messagebox.showerror("Ошибка", f"Не удалось переименовать папку '{old_name}': {e}")

                if renamed_count > 0:
                    messagebox.showinfo("Информация", f"Переименовано папок: {renamed_count}")
                else:
                    messagebox.showinfo("Информация", "Ни одна папка не была переименована.")

                dialog.destroy()
                self.update_tree(sorted_folder_name)

            ttk.Button(dialog, text="Заменить", command=perform_replace).pack(pady=10)

    def write(self, text):
        self.queue.put(text)

    def flush(self):
        pass

    def process_queue(self):
        try:
            while True:
                text = self.queue.get_nowait()
                self.console.insert(tk.END, text)
                self.console.see(tk.END)
        except queue.Empty:
            pass
        finally:
            self.after(100, self.process_queue)

    def on_double_click(self, event):
        item = self.tree.identify('item', event.x, event.y)
        if item:
            self.rename_item(item)

    def rename_item(self, item):
        old_name = self.tree.item(item)['text']
        new_name = simpledialog.askstring("Переименование папки", f"Введите новое имя для папки '{old_name}':")
        if new_name and new_name != old_name:
            parent = self.tree.parent(item)
            if parent:  # Это подпапка
                theme_path = os.path.join(os.path.dirname(__file__), sorted_folder_name, self.tree.item(parent)['text'])
            else:  # Это тема
                theme_path = os.path.join(os.path.dirname(__file__), sorted_folder_name)
            
            old_path = os.path.join(theme_path, old_name)
            new_path = os.path.join(theme_path, new_name)

            try:
                os.rename(old_path, new_path)
                self.tree.item(item, text=new_name)
                self.console.insert(tk.END, f"Папка переименована с '{old_name}' на '{new_name}'\n")
                self.console.see(tk.END)
                
                # Обновляем информацию в self.folder_ids
                folder_uid = self.tree.item(item)['values'][0]
                theme_name, _ = self.folder_ids[folder_uid]
                self.folder_ids[folder_uid] = (theme_name, new_name)
            except OSError as e:
                messagebox.showerror("Ошибка", f"Не удалось переименовать папку: {e}")

    def populate_tree(self, sorted_folder_name):
        self.tree.delete(*self.tree.get_children())
        self.folder_ids = {}
        sorted_path = os.path.join(os.path.dirname(__file__), sorted_folder_name)
        docx_folder = os.path.join(os.path.dirname(__file__), f'docx_{date_range}')
        
        for theme in os.listdir(sorted_path):
            theme_path = os.path.join(sorted_path, theme)
            if os.path.isdir(theme_path):
                theme_uid = str(uuid.uuid4())
                self.folder_ids[theme_uid] = (theme, None)
                theme_id = self.tree.insert('', 'end', text=theme, values=(theme_uid,))
                for subfolder in os.listdir(theme_path):
                    subfolder_path = os.path.join(theme_path, subfolder)
                    if os.path.isdir(subfolder_path):
                        folder_uid = str(uuid.uuid4())
                        self.folder_ids[folder_uid] = (theme, subfolder)
                        
                        docx_file = os.path.join(docx_folder, f"{subfolder}.docx")
                        tags = ('green',) if os.path.exists(docx_file) else ()
                        self.tree.insert(theme_id, 'end', text=subfolder, tags=tags, values=(folder_uid,))

    def update_tree(self, sorted_folder_name):
        # Сохраняем состояние развернутости по тексту элементов
        expanded_items = [self.tree.item(item, 'text') for item in self.tree.get_children() if self.tree.item(item, 'open')]
        
        # Обновляем дерево
        self.populate_tree(sorted_folder_name)
        
        # Восстанавливаем состояние развернутости
        for item in self.tree.get_children():
            if self.tree.item(item, 'text') in expanded_items:
                self.tree.item(item, open=True)

    def merge_selected(self):
        selected_items = self.tree.selection()
        if len(selected_items) < 2:
            messagebox.showwarning("Предупреждение", "Выберите как минимум две папки для объединения.")
            return

        parent = self.tree.parent(selected_items[0])
        if not all(self.tree.parent(item) == parent for item in selected_items):
            messagebox.showwarning("Предупреждение", "Выбранные папки должны находиться в одной теме.")
            return

        new_folder_name = simpledialog.askstring("Объединение папок", "Введите имя новой папки:")
        if not new_folder_name:
            return

        theme_path = os.path.join(os.path.dirname(__file__), sorted_folder_name, self.tree.item(parent)['text'])
        new_folder_path = os.path.join(theme_path, new_folder_name)

        # Если папка уже существует, используем ее, иначе создаем новую
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)

        for item in selected_items:
            folder_name = self.tree.item(item)['text']
            folder_path = os.path.join(theme_path, folder_name)
            if folder_path != new_folder_path:  # Пропускаем папку, если она совпадает с целевой
                for file_name in os.listdir(folder_path):
                    source_file = os.path.join(folder_path, file_name)
                    dest_file = os.path.join(new_folder_path, file_name)
                    try:
                        if not os.path.exists(dest_file):
                            shutil.move(source_file, dest_file)
                        else:
                            # Если файл уже существует, добавляем уникальный суффикс
                            base, ext = os.path.splitext(file_name)
                            counter = 1
                            while os.path.exists(dest_file):
                                dest_file = os.path.join(new_folder_path, f"{base}_{counter}{ext}")
                                counter += 1
                            shutil.move(source_file, dest_file)
                    except PermissionError:
                        print(f"Не удалось переместить файл {source_file}. Возможно, он открыт в другом приложении.")

                # Удаляем пустую папку после перемещения всех файлов
                if not os.listdir(folder_path):
                    os.rmdir(folder_path)
                else:
                    print(f"Папка {folder_path} не пуста и не может быть удалена.")

        self.update_tree(sorted_folder_name)

    def create_docx(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Предупреждение", "Выберите папки или темы для создания DOCX.")
            return

        prefix = simpledialog.askstring("Префикс", "Введите префикс для имен файлов (или оставьте пустым):")

        folders_to_process = []

        for item in selected_items:
            if self.tree.parent(item) == '':  # Это папка темы
                for subitem in self.tree.get_children(item):
                    folder_uid = self.tree.item(subitem)['values'][0]
                    theme_name, folder_name = self.folder_ids[folder_uid]
                    folder_path = os.path.join(os.path.dirname(__file__), sorted_folder_name, theme_name, folder_name)
                    folders_to_process.append((folder_path, folder_name, prefix, folder_uid))
                    self.tree.item(subitem, tags=('gray',))
            else:  # Это подпапка
                folder_uid = self.tree.item(item)['values'][0]
                theme_name, folder_name = self.folder_ids[folder_uid]
                folder_path = os.path.join(os.path.dirname(__file__), sorted_folder_name, theme_name, folder_name)
                folders_to_process.append((folder_path, folder_name, prefix, folder_uid))
                self.tree.item(item, tags=('gray',))

        if not folders_to_process:
            messagebox.showwarning("Предупреждение", "Нет папок для обработки.")
            return

        self.total_tasks += len(folders_to_process)
        self.completed_tasks = 0
        self.progress["maximum"] = 100
        self.update_progress()

        self.open_summary_button.config(state='disabled')
        self.is_processing_docx = True

        for folder_info in folders_to_process:
            self.task_queue.put(folder_info)

        self.tree.selection_remove(self.tree.selection())

        if not self.processing:
            self.process_task_queue()

    def process_task_queue(self):
        if not self.task_queue.empty() and not self.processing:
            self.processing = True
            folder_path, folder_name, prefix, folder_uid = self.task_queue.get()
            threading.Thread(target=self.process_folder, args=(folder_path, folder_name, prefix, folder_uid), daemon=True).start()
        elif self.task_queue.empty() and self.is_processing_docx and self.completed_tasks == self.total_tasks:
            # Все задачи выполнены
            self.is_processing_docx = False
            self.after(0, self.reset_progress)
            self.update_summary_button_state()
        self.after(100, self.process_task_queue)

    def process_folder(self, folder_path, folder_name, prefix, folder_uid):
        try:
            pythoncom.CoInitialize()
            sys.stdout = self
            sys.stderr = self
            
            # Получаем актуальное имя папки из self.folder_ids
            _, actual_folder_name = self.folder_ids[folder_uid]
            actual_folder_path = os.path.join(os.path.dirname(folder_path), actual_folder_name)
            
            print(f"\n--- Начало обработки папки: {actual_folder_name} ---")
            status = process_photos_and_create_docx(actual_folder_path, actual_folder_name, prefix)
            print(f"--- Завершение обработки папки: {actual_folder_name} ---\n")
            
            # Обновляем цвет папки в дереве после создания DOCX
            self.after(0, lambda: self.update_folder_color(folder_uid))
            
            # Обновляем прогресс
            self.completed_tasks += 1
            self.after(0, self.update_progress)
        except Exception as e:
            print(f"Ошибка при обработке папки {actual_folder_name}: {str(e)}")
        finally:
            sys.stdout = sys.__stdout__
            sys.stderr = sys.__stderr__
            pythoncom.CoUninitialize()
            self.processing = False

    def update_folder_color(self, folder_uid):
        for item in self.tree.get_children():
            for subitem in self.tree.get_children(item):
                if self.tree.item(subitem, 'values')[0] == folder_uid:
                    self.tree.item(subitem, tags=('green',))
                    return
            if self.tree.item(item, 'values')[0] == folder_uid:
                self.tree.item(item, tags=('green',))
                return

    def update_progress(self):
        if self.total_tasks > 0:
            percent = int((self.completed_tasks / self.total_tasks) * 100)
            self.progress["value"] = percent
            self.progress_label.config(text=f"Обработка: {percent}%")
            if percent == 100:
                self.progress_label.config(text="Готово")
        else:
            self.progress["value"] = 0
            self.progress_label.config(text="Готово")

    def reset_progress(self):
        self.progress["value"] = 100
        self.progress_label.config(text="Готово")
        self.total_tasks = 0
        self.completed_tasks = 0

    def start_processing(self):
        json_file = os.path.join(os.path.dirname(__file__), 'result.json')
        
        def run_processing():
            self.after(0, lambda: self.update_status("Обновление тем..."))
            update_themes(json_file)
            self.after(0, lambda: self.update_status("Обработка JSON..."))
            process_json(json_file)
            self.after(0, lambda: self.update_status("Готово"))
            self.after(0, lambda: self.populate_tree(sorted_folder_name))

        threading.Thread(target=run_processing, daemon=True).start()

    def update_summary_button_state(self):
        docx_folder = os.path.join(os.path.dirname(__file__), f'docx_{date_range}')
        summary_file = os.path.join(docx_folder, 'Содержание.xlsm')
        if os.path.exists(summary_file) and not self.is_processing_docx and self.completed_tasks == self.total_tasks:
            self.open_summary_button.config(state='normal')
        else:
            self.open_summary_button.config(state='disabled')
        self.after(1000, self.update_summary_button_state)  # Проверяем каждую секунду

    def open_summary(self):
        docx_folder = os.path.join(os.path.dirname(__file__), f'docx_{date_range}')
        summary_file = os.path.join(docx_folder, 'Содержание.xlsm')
        if os.path.exists(summary_file):
            os.startfile(summary_file)
        else:
            messagebox.showwarning("Предупреждение", "Файл содержания не найден.")

    def update_folder_color(self, folder_uid):
        for item in self.tree.get_children():
            for subitem in self.tree.get_children(item):
                if self.tree.item(subitem, 'values')[0] == folder_uid:
                    self.tree.item(subitem, tags=('green',))
                    break

    def update_status(self, message):
        self.status_label.config(text=message)

def main():
    app = Application()
    app.after(100, app.start_processing)
    app.mainloop()

if __name__ == "__main__":
    main()